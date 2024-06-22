#!/usr/bin/env python3
from __future__ import annotations
from typing import Any, Union
from spotipy import Spotify  # type:ignore
from spotipy.oauth2 import SpotifyClientCredentials # type:ignore
from urllib.parse import urlparse
# import json
from tabulate import tabulate, SEPARATING_LINE
from datetime import datetime
import csv
from openpyxl import Workbook
from openpyxl.styles import Font

playlists = [
    "https://open.spotify.com/playlist/3iLEJOL75n0J04oCr15Mhl",  # Fall 2023 - Week 1
    "https://open.spotify.com/playlist/3m06DIJDUizlPFej8FvVGR",
    "https://open.spotify.com/playlist/3m06DIJDUizlPFej8FvVGR",
    "https://open.spotify.com/playlist/1VgvWD4drEeUwSLSzHArYW",
    "https://open.spotify.com/playlist/4ma35bOVN49oLl6xrQ3QAP",
    "https://open.spotify.com/playlist/4ma35bOVN49oLl6xrQ3QAP",
    "https://open.spotify.com/playlist/724NicOVlVZ28177YUXxAY",
    "https://open.spotify.com/playlist/7iiOptIcfDYNerF6eBtQQ2",
    "https://open.spotify.com/playlist/7lJDlXflNmVYcqqFP3o7A5",
    "https://open.spotify.com/playlist/4IuDWNWhkFpCqSDUJPl1WB",
    "https://open.spotify.com/playlist/0dkkmAmpr6mGjNVQDW2jHd",
    "https://open.spotify.com/playlist/7xWOVmfqT3ShCEUE2hsf3n",
    "https://open.spotify.com/playlist/2nLksb3eFyhvSQpV4WiJLk",
    "https://open.spotify.com/playlist/02W29OyFUN8IgnmZt9EWBY",
    "https://open.spotify.com/playlist/2vCOVrT2llplVk12G0XGN7",
    "https://open.spotify.com/playlist/2skkGrmNINM9H6a0odIwmc",   # Spring 2024 - Week 1
    "https://open.spotify.com/playlist/7vMEqna7jz6bVJB6p1h2zi",
    "https://open.spotify.com/playlist/5LiLfykgtptZak0KZqNG7L",
    "https://open.spotify.com/playlist/2HDkPApMLRDXF6IuGrnxtG"
    "https://open.spotify.com/playlist/1NtewVXtntXtn8Ro0nmVG5"
    "https://open.spotify.com/playlist/4pNPkal2JJTQlk7hIBTGy2"
    "https://open.spotify.com/playlist/6YWnpoBROby0SJkAUfGvsI"
    "https://open.spotify.com/playlist/4qfv8ly7r7em7OmucSS8J2"
    "https://open.spotify.com/playlist/64SfZ0trxPNYDKllUL00xU"
    "https://open.spotify.com/playlist/4brtffstgn8dRhf98dVTPr"
]

def get_uri_from_url(url: str) -> str:
    pieces = urlparse(url)
    path = pieces.path
    assert "/playlist/" in path, f"Weird playlist URL: {url}"
    stem = "spotify:playlist:{playlist_id}"
    uri = stem.format(playlist_id=path.split("/")[-1])
    # print(f"Playlist URI: {uri}")
    return uri

def get_id_from_url(url: str) -> str:
    pieces = urlparse(url)
    path = pieces.path
    assert "/playlist/" in path, f"Weird playlist URL: {url}"
    id = path.split("/")[-1]
    # print(f"Playlist ID: {id}")
    return id

def track_to_tuple(track: dict[str, Any]) -> tuple[str, str]:
    artist = ", ".join([artist["name"] for artist in track["artists"]])
    return artist, track["name"]

def get_playlist_info(conn: Spotify, playlist_url: str) -> tuple[str, str]:
    id = get_id_from_url(playlist_url)
    data = conn.playlist(id, fields="name,tracks(items(added_at))")
    # print(json.dumps(data, indent=4, sort_keys=True, ensure_ascii=False))
    assert "tracks" in data, "No tracks data found"
    assert len(data["tracks"]["items"]) > 0, "No individual tracks found"
    name: str = data["name"]
    date: datetime = datetime.strptime(data["tracks"]["items"][0]["added_at"], r"%Y-%m-%dT%H:%M:%SZ")
    # print(date)
    return name, datetime.strftime(date, "%b %Y")

def get_songs(conn: Spotify, playlist_url: str) -> list[tuple[str, str]]:
    """Get songs from a playlist, return a list of artist, track."""
    id = get_id_from_url(playlist_url)
    data: dict[str, Any] = conn.playlist_items(id, fields="items(track(name, artists(name))")
    # print(json.dumps(data, indent=4, sort_keys=True, ensure_ascii=False))
    return [
        track_to_tuple(item["track"])
        for item in data["items"]
    ]

def main() -> None:
    out: list[Union[tuple[str, str], str]] = []

    sp = Spotify(client_credentials_manager=SpotifyClientCredentials())
    with open("vinco-music-league.csv", "w") as fhandle:
        writer = csv.writer(fhandle)
        wb = Workbook()
        ws = wb.active
        assert ws is not None, "No active worksheet?"

        ws.append(["VinCo Music League - complete tracklist"])
        row_id = len(list(ws.rows))
        ws[f"A{row_id}"].font = Font(bold=True, size=24)

        for playlist_url in playlists:
            name, date_added = get_playlist_info(sp, playlist_url)
            tracklist = get_songs(sp, playlist_url)

            # CSV
            writer.writerow([f"{name} - {date_added}", "="*40])
            writer.writerow("")
            for track in tracklist:
                writer.writerow(track)
            writer.writerow("")
            writer.writerow("")

            # XLSX
            # Write title row, bold, merge cells in row
            ws.append([f"{name} - {date_added}",])
            row_id = len(list(ws.rows))
            ws[f"A{row_id}"].font = Font(bold=True)
            ws.append([])
            for track in tracklist:
                ws.append(track)
            ws.append([])

            # Table
            out.append((f"{name} - {date_added}", ""))
            out.append(SEPARATING_LINE)
            out.extend(tracklist)

        artist_colsize = max([len(line[0]) for line in out])
        # 
        track_colsize = max([
            len(line[1]) if line != SEPARATING_LINE else 0
            for line in out
            ])
        ws.column_dimensions['A'].width = artist_colsize
        ws.column_dimensions['B'].width = track_colsize
        wb.save("vinco-music-league.xlsx")

        print(tabulate(out))



if __name__ == "__main__":
    main()
